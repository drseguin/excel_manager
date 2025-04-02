import logging
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string, coordinate_to_tuple
import re

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("excel_manager.log"),
        logging.StreamHandler()
    ]
)

class excelManager:
    def __init__(self, file_path=None):
        """
        Initialize the ExcelManager with an optional file path.
        If no file path is provided, operations will require a file path.
        """
        self.logger = logging.getLogger(__name__)
        self.file_path = file_path
        self.workbook = None
        self.formula_workbook = None
        
        if file_path and os.path.exists(file_path):
            self.load_workbook(file_path)
            self.logger.info(f"Initialized ExcelManager with existing file: {file_path}")
        elif file_path:
            self.create_workbook(file_path)
            self.logger.info(f"Initialized ExcelManager with new file: {file_path}")
        else:
            self.logger.info("Initialized ExcelManager without a file")
    
    def create_workbook(self, file_path=None):
        """
        Create a new Excel workbook.
        """
        path = file_path or self.file_path
        if not path:
            self.logger.error("No file path provided")
            raise ValueError("File path is required to create a workbook")
        
        self.workbook = Workbook()
        # Create a separate workbook for formulas
        self.formula_workbook = Workbook()
        self.file_path = path
        self.save()
        self.logger.info(f"Created new workbook at {path}")
        return self.workbook
    
    def load_workbook(self, file_path=None):
        """
        Load an existing Excel workbook.
        """
        path = file_path or self.file_path
        if not path:
            self.logger.error("No file path provided")
            raise ValueError("File path is required to load a workbook")
        
        if not os.path.exists(path):
            self.logger.error(f"File does not exist: {path}")
            raise FileNotFoundError(f"File does not exist: {path}")
        
        # Load two versions of the workbook - one with formulas and one with calculated values
        self.formula_workbook = load_workbook(path, data_only=False)
        self.workbook = load_workbook(path, data_only=True)
        self.file_path = path
        self.logger.info(f"Loaded workbook from {path}")
        return self.workbook
    
    def save(self, file_path=None):
        """
        Save the workbook to disk.
        """
        if not self.formula_workbook:
            self.logger.error("No workbook loaded")
            raise ValueError("No workbook loaded")
        
        path = file_path or self.file_path
        if not path:
            self.logger.error("No file path provided")
            raise ValueError("File path is required to save a workbook")
        
        # Always save the formula workbook as it contains both formulas and structure
        self.formula_workbook.save(path)
        self.file_path = path
        
        # Reload both workbooks to keep them in sync
        self.formula_workbook = load_workbook(path, data_only=False)
        self.workbook = load_workbook(path, data_only=True)
        
        self.logger.info(f"Saved workbook to {path}")
    
    def close(self):
        """
        Close the workbook.
        """
        if self.workbook:
            self.workbook.close()
            self.workbook = None
        if self.formula_workbook:
            self.formula_workbook.close()
            self.formula_workbook = None
        self.logger.info("Closed workbook")
    
    def count_sheets(self):
        """
        Return the number of sheets in the workbook.
        """
        if not self.formula_workbook:
            self.logger.error("No workbook loaded")
            raise ValueError("No workbook loaded")
        
        count = len(self.formula_workbook.sheetnames)
        self.logger.info(f"Counted {count} sheets")
        return count
    
    def get_sheet_names(self):
        """
        Return the names of the sheets in the workbook.
        """
        if not self.formula_workbook:
            self.logger.error("No workbook loaded")
            raise ValueError("No workbook loaded")
        
        names = self.formula_workbook.sheetnames
        self.logger.info(f"Retrieved sheet names: {names}")
        return names
    
    def create_sheet(self, sheet_name):
        """
        Create a new sheet in the workbook.
        """
        if not self.formula_workbook:
            self.logger.error("No workbook loaded")
            raise ValueError("No workbook loaded")
        
        if sheet_name in self.formula_workbook.sheetnames:
            self.logger.warning(f"Sheet {sheet_name} already exists")
            return self.formula_workbook[sheet_name]
        
        # Create sheet in both workbooks
        formula_sheet = self.formula_workbook.create_sheet(sheet_name)
        value_sheet = self.workbook.create_sheet(sheet_name)
        
        self.logger.info(f"Created new sheet: {sheet_name}")
        return formula_sheet
    
    def get_sheet(self, sheet_name):
        """
        Get a sheet by name.
        """
        if not self.formula_workbook:
            self.logger.error("No workbook loaded")
            raise ValueError("No workbook loaded")
        
        if sheet_name not in self.formula_workbook.sheetnames:
            self.logger.error(f"Sheet does not exist: {sheet_name}")
            raise ValueError(f"Sheet does not exist: {sheet_name}")
        
        formula_sheet = self.formula_workbook[sheet_name]
        self.logger.info(f"Retrieved sheet: {sheet_name}")
        return formula_sheet
    
    def delete_sheet(self, sheet_name):
        """
        Delete a sheet by name.
        """
        if not self.formula_workbook:
            self.logger.error("No workbook loaded")
            raise ValueError("No workbook loaded")
        
        if sheet_name not in self.formula_workbook.sheetnames:
            self.logger.error(f"Sheet does not exist: {sheet_name}")
            raise ValueError(f"Sheet does not exist: {sheet_name}")
        
        # Delete from both workbooks
        del self.formula_workbook[sheet_name]
        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
            
        self.logger.info(f"Deleted sheet: {sheet_name}")
    
    def _parse_cell_reference(self, cell_reference, current_sheet_name=None):
        """
        Parse a cell reference and return the sheet name, row, and column.
        
        Examples:
        - A1: same sheet, row 1, column 1
        - Sheet2!B3: Sheet2, row 3, column 2
        """
        sheet_name = current_sheet_name
        
        # Check if the reference includes a sheet name
        if '!' in cell_reference:
            parts = cell_reference.split('!')
            sheet_name = parts[0].strip("'")
            cell_reference = parts[1]
        
        # Convert A1 reference to row, column
        try:
            column_letter, row = coordinate_from_string(cell_reference)
            column = column_index_from_string(column_letter)
        except Exception as e:
            self.logger.error(f"Invalid cell reference: {cell_reference}. Error: {e}")
            raise ValueError(f"Invalid cell reference: {cell_reference}")
        
        return sheet_name, row, column
    
    def read_cell(self, sheet_name, row_or_cell, column=None):
        """
        Read a cell value. 
        
        Can be called in two ways:
        - read_cell(sheet_name, 'A1') - using cell reference
        - read_cell(sheet_name, 1, 1) - using row and column numbers
        
        Returns the calculated value, not the formula.
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            raise ValueError("No workbook loaded")
        
        # Get the row and column based on the input parameters
        if column is None:
            # Assume row_or_cell is a cell reference like 'A1'
            if not isinstance(row_or_cell, str):
                self.logger.error("Cell reference must be a string")
                raise ValueError("Cell reference must be a string")
            
            sheet_ref, row, col = self._parse_cell_reference(row_or_cell, sheet_name)
            sheet_name = sheet_ref  # Use the sheet name from the reference if provided
        else:
            # Using row and column numbers
            row = row_or_cell
            col = column
        
        if sheet_name not in self.workbook.sheetnames:
            self.logger.error(f"Sheet does not exist: {sheet_name}")
            raise ValueError(f"Sheet does not exist: {sheet_name}")
        
        # Get the calculated value from the data_only workbook
        sheet = self.workbook[sheet_name]
        value = sheet.cell(row=row, column=col).value
        
        if value is None:
            value = ''
        elif isinstance(value, float):
            value = round(value, 2)

        # Check if cell is formatted as currency and prepend '$'
        formula_cell = self.formula_workbook[sheet_name].cell(row=row, column=col)
        if formula_cell.number_format and '$' in formula_cell.number_format and isinstance(value, (int, float)):
            value = "$" + str(value)

        # Get the formula (if any) from the formula workbook for logging
        formula = formula_cell.value
        
        cell_ref = f"{get_column_letter(col)}{row}"
        if isinstance(formula, str) and formula.startswith('='):
            self.logger.info(f"Read calculated value '{value}' from cell {cell_ref} in sheet {sheet_name} (formula: {formula})")
        else:
            self.logger.info(f"Read value '{value}' from cell {cell_ref} in sheet {sheet_name}")
        
        return value
    
    def write_cell(self, sheet_name, row_or_cell, column=None, value=None):
        """
        Write a value to a cell.
        
        Can be called in two ways:
        - write_cell(sheet_name, 'A1', value) - using cell reference
        - write_cell(sheet_name, 1, 1, value) - using row and column numbers
        """
        if not self.formula_workbook:
            self.logger.error("No workbook loaded")
            raise ValueError("No workbook loaded")
        
        # Get the row and column based on the input parameters
        if column is None and value is None:
            self.logger.error("Missing value parameter")
            raise ValueError("Missing value parameter")
        elif column is not None and value is None:
            # Assume row_or_cell is a cell reference and column is the value
            value = column
            if not isinstance(row_or_cell, str):
                self.logger.error("Cell reference must be a string")
                raise ValueError("Cell reference must be a string")
            
            sheet_ref, row, col = self._parse_cell_reference(row_or_cell, sheet_name)
            sheet_name = sheet_ref  # Use the sheet name from the reference if provided
        else:
            # Using row and column numbers
            row = row_or_cell
            col = column
        
        if sheet_name not in self.formula_workbook.sheetnames:
            self.logger.error(f"Sheet does not exist: {sheet_name}")
            raise ValueError(f"Sheet does not exist: {sheet_name}")
        
        # Write to the formula workbook
        formula_sheet = self.formula_workbook[sheet_name]
        formula_sheet.cell(row=row, column=col).value = value
        
        cell_ref = f"{get_column_letter(col)}{row}"
        self.logger.info(f"Wrote value '{value}' to cell {cell_ref} in sheet {sheet_name}")
    
    def read_range(self, sheet_name, start_cell_or_row, start_column=None, end_cell_or_row=None, end_column=None):
        """
        Read a range of cells.
        
        Can be called in two ways:
        - read_range(sheet_name, 'A1:C3') - using cell range reference
        - read_range(sheet_name, 'A1', 'C3') - using start and end cell references
        - read_range(sheet_name, 1, 1, 3, 3) - using row and column numbers
        
        Returns the calculated values, not the formulas.
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            raise ValueError("No workbook loaded")
        
        # Parse the arguments to determine start and end coordinates
        if isinstance(start_cell_or_row, str) and ':' in start_cell_or_row and start_column is None:
            # Range notation like 'A1:C3'
            start_ref, end_ref = start_cell_or_row.split(':')
            sheet_ref, start_row, start_col = self._parse_cell_reference(start_ref, sheet_name)
            _, end_row, end_col = self._parse_cell_reference(end_ref, sheet_name)
            sheet_name = sheet_ref  # Use the sheet name from the reference if provided
        elif isinstance(start_cell_or_row, str) and isinstance(start_column, str) and end_cell_or_row is None:
            # Two cell references like 'A1', 'C3'
            sheet_ref, start_row, start_col = self._parse_cell_reference(start_cell_or_row, sheet_name)
            _, end_row, end_col = self._parse_cell_reference(start_column, sheet_name)
            sheet_name = sheet_ref  # Use the sheet name from the reference if provided
        elif all(param is not None for param in [start_column, end_cell_or_row, end_column]):
            # Row and column numbers
            start_row = start_cell_or_row
            start_col = start_column
            end_row = end_cell_or_row
            end_col = end_column
        else:
            self.logger.error("Invalid arguments for read_range")
            raise ValueError("Invalid arguments for read_range")
        
        if sheet_name not in self.workbook.sheetnames:
            self.logger.error(f"Sheet does not exist: {sheet_name}")
            raise ValueError(f"Sheet does not exist: {sheet_name}")
        
        # Get the calculated values from the data_only workbook
        sheet = self.workbook[sheet_name]
        values = []
        for row in range(start_row, end_row + 1):
            row_values = []
            for col in range(start_col, end_col + 1):
                cell_val = sheet.cell(row=row, column=col).value
                if cell_val is None:
                    cell_val = ''
                elif isinstance(cell_val, float):
                    cell_val = round(cell_val, 2)

                # Check if cell is formatted as currency and prepend '$'
                formula_cell = self.formula_workbook[sheet_name].cell(row=row, column=col)
                if formula_cell.number_format and '$' in formula_cell.number_format and cell_val is not None and isinstance(cell_val, (int, float)):
                    cell_val = "$" + str(cell_val)
                row_values.append(cell_val)
            values.append(row_values)
        
        range_ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        self.logger.info(f"Read range {range_ref} in sheet {sheet_name}")
        return values
    
    def write_range(self, sheet_name, start_cell_or_row, start_column_or_values=None, values_or_end_row=None, end_column=None):
        """
        Write values to a range of cells.
        
        Can be called in three ways:
        - write_range(sheet_name, 'A1', values) - using cell reference for start
        - write_range(sheet_name, 1, 1, values) - using row and column numbers for start
        """
        if not self.formula_workbook:
            self.logger.error("No workbook loaded")
            raise ValueError("No workbook loaded")
        
        # Parse the arguments to determine the start coordinate and values
        if isinstance(start_cell_or_row, str) and values_or_end_row is None:
            # Cell reference like 'A1' and values
            values = start_column_or_values
            sheet_ref, start_row, start_col = self._parse_cell_reference(start_cell_or_row, sheet_name)
            sheet_name = sheet_ref  # Use the sheet name from the reference if provided
        elif values_or_end_row is not None and end_column is None:
            # Row and column numbers for start, and values
            start_row = start_cell_or_row
            start_col = start_column_or_values
            values = values_or_end_row
        else:
            self.logger.error("Invalid arguments for write_range")
            raise ValueError("Invalid arguments for write_range")
        
        if sheet_name not in self.formula_workbook.sheetnames:
            self.logger.error(f"Sheet does not exist: {sheet_name}")
            raise ValueError(f"Sheet does not exist: {sheet_name}")
        
        # Write to the formula workbook
        formula_sheet = self.formula_workbook[sheet_name]
        for i, row_values in enumerate(values):
            for j, value in enumerate(row_values):
                formula_sheet.cell(row=start_row + i, column=start_col + j).value = value
        
        end_row = start_row + len(values) - 1
        end_col = start_col + len(values[0]) - 1 if values else start_col
        range_ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        
        self.logger.info(f"Wrote values to range {range_ref} in sheet {sheet_name}")
        
    def read_total(self, sheet_name, row_or_cell, column=None):
        """
        Read the total value by traversing down rows until an empty cell is found.
        Then back up one cell and return that value.
        
        Can be called in two ways:
        - read_total(sheet_name, 'A1') - using cell reference
        - read_total(sheet_name, 1, 1) - using row and column numbers
        
        Returns the calculated total value, typically found at the end of a column of values.
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            raise ValueError("No workbook loaded")
        
        # Get the row and column based on the input parameters
        if column is None:
            # Assume row_or_cell is a cell reference like 'A1'
            if not isinstance(row_or_cell, str):
                self.logger.error("Cell reference must be a string")
                raise ValueError("Cell reference must be a string")
            
            sheet_ref, start_row, start_col = self._parse_cell_reference(row_or_cell, sheet_name)
            sheet_name = sheet_ref  # Use the sheet name from the reference if provided
        else:
            # Using row and column numbers
            start_row = row_or_cell
            start_col = column
        
        if sheet_name not in self.workbook.sheetnames:
            self.logger.error(f"Sheet does not exist: {sheet_name}")
            raise ValueError(f"Sheet does not exist: {sheet_name}")
        
        # Get the sheet from the data_only workbook to read calculated values
        sheet = self.workbook[sheet_name]
        
        # Start from the given cell and traverse down
        current_row = start_row
        max_rows = sheet.max_row
        
        # Keep track of the last non-empty cell value encountered
        last_value = None
        last_row = None
        
        while current_row <= max_rows:
            value = sheet.cell(row=current_row, column=start_col).value
            
            # If we find an empty cell and we've seen at least one non-empty cell, 
            # we'll return the last non-empty cell value (which should be the total)
            if value is None or value == '':
                if last_value is not None:
                    # Format the value if needed
                    if isinstance(last_value, float):
                        last_value = round(last_value, 2)
                    
                    # Check if cell is formatted as currency and prepend '$'
                    formula_cell = self.formula_workbook[sheet_name].cell(row=last_row, column=start_col)
                    if formula_cell.number_format and '$' in formula_cell.number_format and isinstance(last_value, (int, float)):
                        last_value = "$" + str(last_value)
                    
                    cell_ref = f"{get_column_letter(start_col)}{last_row}"
                    self.logger.info(f"Found total value '{last_value}' at cell {cell_ref} in sheet {sheet_name}")
                    return last_value
                else:
                    # If we haven't found any non-empty cells, continue searching
                    current_row += 1
                    continue
            
            # Update the last non-empty value seen
            last_value = value
            last_row = current_row
            current_row += 1
        
        # If we reach the end of the sheet and have a value, return it
        if last_value is not None:
            # Format the value if needed
            if isinstance(last_value, float):
                last_value = round(last_value, 2)
            
            # Check if cell is formatted as currency and prepend '$'
            formula_cell = self.formula_workbook[sheet_name].cell(row=last_row, column=start_col)
            if formula_cell.number_format and '$' in formula_cell.number_format and isinstance(last_value, (int, float)):
                last_value = "$" + str(last_value)
            
            cell_ref = f"{get_column_letter(start_col)}{last_row}"
            self.logger.info(f"Found total value '{last_value}' at cell {cell_ref} in sheet {sheet_name} (at end of sheet)")
            return last_value
        
        # If no non-empty cells were found
        self.logger.warning(f"No values found starting from {get_column_letter(start_col)}{start_row} in sheet {sheet_name}")
        return None